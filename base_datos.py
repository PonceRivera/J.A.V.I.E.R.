import pandas as pd
import os

def cargar_inventario():
    ruta_excel = os.path.join(os.path.dirname(__file__), 'Nuevo inventario 2026.xlsx')
    try:
        # Usamos header=1 porque la primera fila del excel está vacía o tiene un título
        return pd.read_excel(ruta_excel, header=1)
    except FileNotFoundError:
        print("Error: No se encontró el archivo Nuevo inventario 2026.xlsx")
        return None

def buscar_articulo(nombre_buscado):
    """
    Busca un artículo en el Excel y retorna si hay disponibles.
    """
    df = cargar_inventario()
    if df is None:
        return "Error al leer la base de datos del inventario."
    
    nombre_buscado = nombre_buscado.lower()
    df = df.fillna('') # Para que no truene si hay celdas vacías
    
    for index, row in df.iterrows():
        articulo = str(row.get('Herramienta', '')).lower()
        if articulo == '':
            continue
            
        if nombre_buscado in articulo or articulo in nombre_buscado:
            cantidad = row.get('cantidad ', row.get('cantidad', 'una cantidad desconocida'))
            # En este excel de enero no hay columna de "Ubicación", así que le decimos la cantidad
            return f"Sí tenemos {row.get('Herramienta')}. Según el inventario de enero, hay {cantidad} disponibles."
            
    return f"Lo siento, no encontré {nombre_buscado} en el inventario de enero."

def listar_inventario_completo():
    """
    Retorna una lista de todos los artículos en el inventario con sus cantidades.
    """
    df = cargar_inventario()
    if df is None:
        return "Error al leer la base de datos del inventario."
    
    df = df.fillna('')
    items = []
    for _, row in df.iterrows():
        articulo = str(row.get('Herramienta', '')).strip()
        if articulo == '':
            continue
        cantidad = row.get('cantidad ', row.get('cantidad', '?'))
        items.append(f"{articulo}: {cantidad}")
    
    if not items:
        return "El inventario está vacío o no pude leerlo correctamente."
    
    total = len(items)
    resumen = f"El inventario tiene {total} artículos en total. Aquí están: " + ". ".join(items) + "."
    return resumen

def buscar_inventario_por_categoria(categoria: str):
    """
    Filtra el inventario por una categoría o palabra clave (ej. 'electrónica', 'tornillos').
    """
    df = cargar_inventario()
    if df is None:
        return "Error al leer la base de datos del inventario."
    
    df = df.fillna('')
    items = []
    cat = categoria.lower()
    for _, row in df.iterrows():
        articulo = str(row.get('Herramienta', '')).strip()
        if articulo == '':
            continue
        # Buscar por nombre de herramienta o columna de categoría si existe
        categoria_col = str(row.get('Categoria', row.get('Categoría', row.get('Tipo', '')))).lower()
        if cat in articulo.lower() or cat in categoria_col:
            cantidad = row.get('cantidad ', row.get('cantidad', '?'))
            items.append(f"{articulo}: {cantidad}")
    
    if not items:
        return f"No encontré artículos relacionados con '{categoria}' en el inventario."
    
    total = len(items)
    return f"Encontré {total} artículos relacionados con '{categoria}': " + ". ".join(items) + "."

def actualizar_inventario(nombre_articulo: str, cantidad: str):
    """
    Agrega o actualiza la cantidad de un artículo en el inventario.
    """
    ruta_excel = os.path.join(os.path.dirname(__file__), 'Nuevo inventario 2026.xlsx')
    try:
        from openpyxl import load_workbook
        wb = load_workbook(ruta_excel)
        ws = wb.active
        
        encontrado = False
        for row in range(3, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and str(cell_val).strip().lower() == nombre_articulo.lower():
                ws.cell(row=row, column=2).value = cantidad
                encontrado = True
                break
                
        if not encontrado:
            ws.append([nombre_articulo, cantidad, "", "", ""])
            
        wb.save(ruta_excel)
        return f"Inventario actualizado exitosamente. '{nombre_articulo}' ahora tiene la cantidad de '{cantidad}'."
    except Exception as e:
        return f"Error al actualizar el inventario: {e}"

if __name__ == "__main__":
    # Prueba rápida
    print(buscar_articulo("pinzas"))
    print(listar_inventario_completo())
